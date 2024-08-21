import networkx as nx
import pandas as pd
import time
import numpy as np
import random
import itertools

import openpyxl
from openpyxl import load_workbook

# ************Excel deney parametre okuma başlangıç
# Excel dosyasının yolu
dosya_yolu = "DeneylerPath.xlsx"
# Okunacak sayfa adı
sayfa_adi = "Makale3"


def oku_excel_satirlar_sutunlar_ile(excel_dosya, sayfa):
    # Excel dosyasını aç
    workbook = openpyxl.load_workbook(excel_dosya)
    # Belirtilen sayfayı seç
    sheet = workbook[sayfa]

    # Sütun başlıklarını al
    sütun_başlıkları = [hucre.value for hucre in sheet[1]]

    # Tüm satırları ve sütun değerlerini tutacak ana liste
    satir_sutun_degerleri = []

    # Her bir satırı dolaş
    for satir in sheet.iter_rows(min_row=2, values_only=True):
        # Her bir sütun değerini, sütun başlığı ile birlikte bir sözlükte tut
        satir_sutun_dict = {}
        for index, deger in enumerate(satir):
            satir_sutun_dict[sütun_başlıkları[index]] = deger
        # Her bir satırın sütun değerlerini ana listeye ekle
        satir_sutun_degerleri.append(satir_sutun_dict)

    return satir_sutun_degerleri


# Excel dosyasını oku ve satır-sütun değerlerini birer değişkene atayarak al
satir_sutun_degerleri = oku_excel_satirlar_sutunlar_ile(dosya_yolu, sayfa_adi)

kimbinasyonsayisi = 10
secimsayisi = [5,10,20,40,50]  # seçim sayısı
islemtekrarsayisi = 10
dugumderecesi = ""
motiforani = ""


# ************Excelparametre okuma bitiş
def sonuclandir(dugumpath, motifpath, dugumderecesi, motiforani,motiftipi,nodepath,mptifpath):
    # Mevcut Excel sonuclar dosyasını yükle
    file_path = 'DeneySonuclari.xlsx'

    # dugumlerkonum="Kaynaklar/Deney1/graph_500Nodes_1Layers3AvgDegree1MotifType0.300000MutationRate1GraphType.txt"
    # motiflerkonum="Kaynaklar/Deney1/Motif_graph_500Nodes_1Layers3AvgDegree1MotifType0.300000MutationRate1GraphType.txt"
    dugumlerkonum = "mn/d3/" + dugumpath
    motiflerkonum = "mn/d3/" + motifpath  # "Motif_"+str(dugumpath)

    kimbinasyonsayisi = 10
    secimsayisi = [5,10,20,40,50]  # seçim sayısı
    islemtekrarsayisi = 10
    dugumderecesi = dugumderecesi
    motiforani = motiforani

    # kimbinasyonsayisi = 3
    # secimsayisi = [1, 5, 10, 20, 40]  # seçim sayısı
    # islemtekrarsayisi = 3
    # dugumderecesi=3
    # motiforani=40
    # motifsayisi=500 verilen networkten okunup atanıyor
    # dugumsayisi=100 verilen networkten okunup atanıyor
    # algoritma=greedy1  calisan algoritma ismi atanıyor
    def add_row_to_excel(file_path, data_dict):
        """
        Verilen Excel dosyasına yeni bir satır ekler.

        Args:
        file_path (str): Excel dosyasının yolu.
        data_dict (dict): Sütun başlıkları ile karşılık gelen değerlerin dictionary'i.
                          Örneğin, {'Algoritma': 'Greedy1', 'dugumsayisi': 500, 'dugumderecesi': 3, 'motiforani':40, ''}
        """
        # Excel dosyasını yükle
        wb = load_workbook(file_path)
        ws = wb.active  # Aktif çalışma sayfasını seç

        # Sütun başlıklarını al
        header_row = {cell.value: cell.column for cell in ws[1] if cell.value}

        # Yeni satırı oluştur (varsayılan olarak None ile doldurulur)
        new_row = [None] * ws.max_column

        # Verilen data_dict içindeki verileri uygun sütunlara yerleştir
        for key, value in data_dict.items():
            if key in header_row:
                # Sütun numarasını bul ve yeni satırdaki doğru pozisyona değeri yerleştir
                col_index = header_row[key] - 1  # Sütun indeksi 1'den başlar, liste indeksi 0'dan
                new_row[col_index] = value

        # Yeni satırı worksheet'e ekle
        ws.append(new_row)

        # Dosyayı kaydet
        wb.save(file_path)
        # print("Yeni satır başarıyla eklendi.")

    # excele kayıt metodu örnek kullanım
    # data_dict={'algoritma':'greedy1','dugumsayisi':500,'dugumderecesi':3,'motiforani':40,'secilenmotifsayisi':1,'motifsayisi':103,'mesafe':24534,'zamansn':0.001654}
    # add_row_to_excel(file_path,data_dict)

    G = nx.read_adjlist(dugumlerkonum)  # yönlendirilmişler için DiGraph

    print(len(G.edges))
    # ***************Kopuk Düğüm ve Graphları networke dahil etme başlanıgıcı*****************************
    connected_components = list(nx.connected_components(G))

    # Ana ağı ve alt ağları tespit
    main_component = max(connected_components, key=len)
    sub_networks = [comp for comp in connected_components if comp != main_component]

    # Ana ağın bir düğümünü seçiyoruz
    main_node = next(iter(main_component))

    # Alt ağları ana ağa bağlayacak kenarları ekliyoruz
    for sub_network in sub_networks:
        sub_node = next(iter(sub_network))
        G.add_edge(main_node, sub_node)

    # ***************Kopuk Düğüm ve Graphları networke dahil etme bitiş*****************************
    print(len(G.edges))

    # exit()

    node_node_distance_matris = []
    for nodei in list(G.nodes):
        row = []
        rowsum = 0
        for nodej in list(G.nodes):
            if nx.has_path(G, source=nodei, target=nodej):
                length = nx.shortest_path_length(G, source=nodei, target=nodej)
                rowsum = rowsum + length
            else:
                length = len(G.nodes)
            row.append(length)
        # row.append(rowsum)  # her satır sonundaki son sütuna o satır elemanlarının toplamı ekleniyor. sütun sayısı toplam düğüm sayısı+1 dir
        node_node_distance_matris.append(row)

    dugumsayisi = len(G.nodes)  # excele kaydetmek için
    print("Düğümler Okundu")
    # motiflerin formatlanması

    motifler = []
    with open(motiflerkonum, 'r') as dosya:
        for satir in dosya:
            temiz_satir = satir.strip().replace(" ", "").replace(")(", ",").replace("(", "").replace(")", "")
            koordinatlar = temiz_satir.split(",")
            motifdizisi = []
            for koordinat in koordinatlar:
                motifdizisi.append(str(koordinat.strip()))
            # Koordinat dizisini ana listeye ekle
            motifler.append(motifdizisi)

    # bir sütuna göre en küçük değere sahip satırı döndürür
    def en_yakin_motif(matrix, sutun_indeksi):
        en_kucuk_deger = float('inf')
        en_kisa_motif = None
        for row in matrix:
            if row[sutun_indeksi] < en_kucuk_deger:
                en_kucuk_deger = row[sutun_indeksi]
                en_kisa_motif = row
        return en_kisa_motif

    motif_node_distance_matris = []
    for motif in list(motifler):
        motif_node_distance_submatris = []
        for nodei in list(set(motif)):
            row = []
            # rowsum = 0
            subrow = []
            # subrow.append(motif)
            length = len(G.nodes)
            for nodej in list(G.nodes):
                try:
                    if nx.has_path(G, source=nodei, target=nodej):
                        length = nx.shortest_path_length(G, source=nodei, target=nodej)
                        # rowsum = rowsum + length
                    # else:
                    #    length = len(G.nodes)  # kenar bağlantısı olmayan düğümlere max uzunluk değeri veriliyor
                except:
                    length = len(G.nodes)
                    # print(nodei + ", " + nodej)
                subrow.append(length)
            # subrow.append(nodei)
            # subrow.append(rowsum)
            motif_node_distance_submatris.append(subrow)

        eym = en_yakin_motif(motif_node_distance_submatris, len(motif_node_distance_submatris))
        motif_node_distance_matris.append(eym)

    motifsayisi = len(motifler)  # excele kaydetmek için
    print("Motif Düğüm Mesafeleri Hesaplandı :" + str(motifsayisi))

    # İşlenecek motif node matris
    matrix = np.array(motif_node_distance_matris)

    """
    # Örnek Matris
    matrix = np.array([[1, 3, 2, 4, 4, 7],
                       [3, 1, 5, 4, 2, 3],
                       [2, 5, 1, 1, 4, 3],
                       [5, 7, 1, 3, 1, 1],
                       [2, 3, 5, 6, 7, 7],
                       [1, 2, 5, 6, 7, 2]])
    """
    """
    En küçük satırlar hesaplama fonk baslangic
    """

    def en_kucuk_sutun_degerleri_ve_toplami(matris):
        # Matrisin sütun sayısını al
        sutun_sayisi = len(matris[0])

        # Yeni bir liste oluştur, bu listede her bir sütundaki en küçük değeri ve toplamını tutacağız
        en_kucukler = []
        toplam = 0

        # Her bir sütunu dolaşarak en küçük değerleri bul ve toplamlarını hesapla
        for j in range(sutun_sayisi):
            # İlgili sütunun değerlerini al ve en küçük değeri bul
            en_kucuk = min(matris[i][j] for i in range(len(matris)))

            # Bulduğumuz en küçük değeri en_kucukler listesine ekle
            en_kucukler.append(en_kucuk)

            # Sütun değerlerini topla
            toplam += en_kucuk

        # Sonuç listesini ve sütun değerleri toplamını döndür
        return en_kucukler, toplam

    """
    En küçük satırlar hesaplama fonk bitis
    """

    """
    Greedy1 başlangıç
    """

    def greedy1alg(matrix, k):
        start_time = time.time()  # İşlem başlangıç zamanı

        original_matrix = matrix.copy()
        original_indices = np.arange(len(matrix))  # Orijinal matrisin indekslerini sakla
        removed_rows = []
        removed_original_rows = []  # Çıkarılan satırların orijinal değerlerini tutacak liste
        removed_original_indices = []  # Çıkarılan satırların orijinal indekslerini tutacak liste

        # start_time = time.time()  # İşlem başlangıç zamanı

        for _ in range(min(k, matrix.shape[0])):
            # Sütun toplamlarını hesapla
            column_sums = np.sum(matrix, axis=0)

            # En küçük toplama sahip satırı bul
            min_row_index = np.argmin(np.sum(matrix, axis=1))
            removed_row = matrix[min_row_index, :]
            removed_original_row = original_matrix[min_row_index, :]  # Orijinal sütun değerlerini al
            removed_original_index = original_indices[min_row_index]  # Orijinal indeksi al

            # Çıkarılan satırın orijinal değerlerini ve indeksini kaydet
            removed_original_rows.append(removed_original_row)
            removed_original_indices.append(removed_original_index)

            # Matristen ve orijinal matristen satırı çıkar
            matrix = np.delete(matrix, min_row_index, axis=0)
            original_matrix = np.delete(original_matrix, min_row_index, axis=0)
            original_indices = np.delete(original_indices, min_row_index)  # İndis listesinden de çıkar

            if matrix.size == 0:  # Eğer matris boşsa döngüyü bitir
                removed_rows.append(np.sum(removed_row))
                break

            # Çıkarılan satırın sütun değerlerini kalan satırlarla karşılaştır ve minimum değerleri bul
            for i in range(matrix.shape[0]):
                matrix[i] = np.minimum(matrix[i], removed_row)

            # Çıkarılan satırın toplamını (orijinal değerlerle) kaydet
            removed_rows.append(np.sum(removed_row))

        # Toplam değerlere göre sırala
        sorted_indices = np.argsort(removed_rows)
        sorted_indices = sorted_indices[::-1]  # Ters sıralama
        sorted_removed_rows = [removed_original_rows[i] for i in sorted_indices]
        sorted_removed_indices = [removed_original_indices[i] for i in sorted_indices]

        end_time = time.time()  # İşlem bitiş zamanı
        elapsed_time = end_time - start_time  # İşlem süresi

        # bileşke satır ve satır toplamı yazdırma başlangıç
        # seçilen satırların matrise aktarılması
        yeni_satir_matrisi = []
        print("Test Satırlar")
        for item in enumerate(sorted_removed_rows):
            yeni_satir_matrisi.append(item[1])

        # print(yeni_satir_matrisi)
        # matrise ait en küçük sütun değerli satırın oluşturulması
        satir = en_kucuk_sutun_degerleri_ve_toplami(yeni_satir_matrisi)
        print(satir[0])
        print(satir[1])
        print("************")
        # print("Toplam: " + str(np.sum(yeni_satir_matrisi[0], axis=0)))
        # Bileşke satır ve toplamı yazdırma bitiş

        data_dict = {'algoritma': 'greedy1', 'dugumsayisi': dugumsayisi, 'dugumderecesi': dugumderecesi,
                     'motiforani': motiforani, 'motifsayisi': motifsayisi, 'k': k, 'mesafe': satir[1],
                     'zamansn': elapsed_time, 'motiftipi': motiftipi, 'nodepath':nodepath, 'motifpath':motifpath}
        add_row_to_excel(file_path, data_dict)

        # return en_kucuk_deger, sorted_removed_indices, removed_rows, elapsed_time  #tüm birleşimi oluşturan satırlar ve değerlerini döndürür
        return sorted_removed_rows, sorted_removed_indices, removed_rows, elapsed_time

    """
    Greedy1 bitiş
    """

    """
    Greedy2 Başlangıç
    """

    def greedy2alg(matrix, k):
        start_time = time.time()  # Zaman ölçümünü başlat

        original_matrix = np.array(matrix)
        rows_order = []
        original_indices = list(range(original_matrix.shape[0]))
        union_min_values = np.inf * np.ones(original_matrix.shape[1])

        while len(rows_order) < k and original_matrix.size > 0:
            # En küçük değeri bul
            min_value = np.min(original_matrix)

            # Her satırdaki en küçük değerin sayısını bul
            counts_per_row = (original_matrix == min_value).sum(axis=1)

            # En fazla bu değere sahip sütunun olduğu satırı bul
            row_with_most_min = np.argmax(counts_per_row)

            # Çıkarılan satırın orijinal halini ve indexini sakla
            original_row_index = original_indices.pop(row_with_most_min)
            rows_order.append((original_row_index, matrix[original_row_index]))

            # Çıkarılan satırdaki en küçük değerleri birleşim için kaydet
            union_min_values = np.minimum(union_min_values, matrix[original_row_index])

            # Matristen satırı çıkar
            original_matrix = np.delete(original_matrix, row_with_most_min, 0)

            # Eğer matris boşsa döngüyü bitir
            if original_matrix.size == 0:
                break

            # Çıkarılan satırdaki en küçük değerlerin indekslerini bul
            cols_with_min = np.where(rows_order[-1][1] == min_value)[0]

            # Kalan satırların sütunlarını güncelle
            for col in cols_with_min:
                original_matrix[:, col] = min_value

        # Çıkarılan satırları ve sütun toplamlarını yazdır
        # for index, (row_index, row_values) in enumerate(rows_order):
        #    column_sum = np.sum(row_values)  # Çıkarılan satırın sütun toplamını hesapla
        #    # print(f"Çıkarılma Sırası: {index}, Orijinal Satır Indexi: {row_index}, Satır Değerleri: {row_values}, Sütun Toplamı: {column_sum}")
        #    print(f"Çıkarılma Sırası: {index}, Orijinal Satır Indexi: {row_index},  Sütun Toplamı: {column_sum}")

        yeni_satir_matrisi = []
        print("Satırlar")
        for index, (row_index, row_values) in enumerate(rows_order):
            yeni_satir_matrisi.append(row_values)
        # print(yeni_satir_matrisi)
        print("birleşim satırı")
        satir = en_kucuk_sutun_degerleri_ve_toplami(yeni_satir_matrisi)
        print(satir[0])  # birleşim satırı değerleri
        print(satir[1])  # birleşim satırı toplamı
        print("******************")
        # Çıkarılan satırların sütunlarındaki en küçük değerlerin birleşimini yazdır
        # print(f"Çıkarılan satırların sütunlarındaki en küçük değerlerin birleşimi: {union_min_values}")

        # Çalışma süresini yazdır
        end_time = time.time()  # Zaman ölçümünü bitir
        elapsed_time = end_time - start_time

        data_dict = {'algoritma': 'greedy2', 'dugumsayisi': dugumsayisi, 'dugumderecesi': dugumderecesi,
                     'motiforani': motiforani, 'motifsayisi': motifsayisi, 'k': k, 'mesafe': satir[1],
                     'zamansn': elapsed_time}
        add_row_to_excel(file_path, data_dict)

        print(f"Kodun çalışma süresi: {end_time - start_time:.6f} saniye")

    """
    Greedy2 bitiş
    """
    """
    Random Başlangıç
    """

    def randomalg(matrix, n, k):
        start_time = time.time()  # Zaman ölçümünü başlat

        np_matrix = np.array(matrix)
        union_rows = []
        all_selected_rows = []

        for _ in range(k):
            # Rastgele n satır seç ve indekslerini kaydet
            selected_indices = random.sample(range(np_matrix.shape[0]), n)
            selected_rows = np_matrix[selected_indices]
            all_selected_rows.append(selected_indices)

            # Seçilen satırların her sütunu için en küçük değerleri bul
            min_values = selected_rows.min(axis=0)
            union_rows.append((min_values, sum(min_values)))  # Birleşim ve toplamını kaydet
        """
        print(n)
        print(k)
        print(len(all_selected_rows))
        print("Bilgi" + str(all_selected_rows))
        for itemss in all_selected_rows:
            print(itemss)
        """
        # Birleşim satırlarını toplamlarına göre küçükten büyüğe sırala
        union_rows_sorted = sorted(union_rows, key=lambda x: x[1])
        """
        # Seçilen satırları ve birleşimleri yazdır
        for i, (min_values, total) in enumerate(union_rows_sorted):
            print(f"Kombinasyon {i + 1}:")
            #print(f"Seçilen satırların indeksleri: {all_selected_rows[i]}")
            print(f"Seçilen satırların değerleri:")
            #for index in all_selected_rows[i]:
            #    print(f"Satır {index}: {matrix[index]}")
            #print(f"En küçük değerlerin birleşimi: {min_values}")
            print(f"Birleşim satırının toplamı: {total}")
            print()
        """
        #standart sapma hesaplama başlangıc

        # Belirtilen indislerdeki satırları seç
        secilen_satirlar = matrix[selected_indices, :]
        # Seçilen satırların sütun toplamlarını hesapla
        sutun_toplamlari = np.sum(secilen_satirlar, axis=0)

        # Sütun toplamlarının standart sapmasını hesapla
        standart_sapma = np.std(sutun_toplamlari)

        print("Seçilen satırların sütun toplamlarının standart sapması:", standart_sapma)

        # standart sapma hesaplama Bitiş

        toplam = 0
        for i, (min_values, total) in enumerate(union_rows_sorted):
            toplam = toplam + total

        toplam = toplam / k
        # Toplam işlem süresini yazdır
        elapsed_time = time.time() - start_time

        data_dict = {'algoritma': 'random', 'dugumsayisi': dugumsayisi, 'dugumderecesi': dugumderecesi,
                     'motiforani': motiforani, 'motifsayisi': motifsayisi, 'k': n, 'mesafe': toplam,
                     'zamansn': elapsed_time, 'ssapma': standart_sapma, 'motiftipi': motiftipi,'nodepath':nodepath, 'motifpath':motifpath}
        add_row_to_excel(file_path, data_dict)

        print(f"Toplam işlem süresi: {elapsed_time:.6f} saniye")

    """
    Random bitiş
    """
    """
    BruteForce Başlangıç
    """

    """
    def bruteforcealg(matris, kn, k):
        start_time = time.time()  # İşlem başlangıç zamanı
        # Tüm satırların olası k'lı kombinasyonlarını oluştur
        kombinasyonlar = list(itertools.combinations(range(len(matris)), kn))

        # Her kombinasyon için sütun değerlerinin minimumlarını ve toplamlarını hesapla
        kombinasyon_degerleri = []


        kmatris=[]
        # combinations('ABCD', 2) → AB AC AD BC BD CD
        # combinations(range(4), 3) → 012 013 023 123
        pool = tuple(matris)
        n = len(pool)
        r=kn
        if r > n:
            return
        indices = list(range(r))
        # yield tuple(pool[i] for i in indices)
        print(str(tuple(pool[i] for i in indices)))

        satir = en_kucuk_sutun_degerleri_ve_toplami(tuple(pool[i] for i in indices))
        print(satir[0])
        print(satir[1])
        print("ilk force")
        exit()
        kmatris.append(satir[1], tuple(pool[i] for i in indices), satir[0])
        while True:
            for i in reversed(range(r)):
                if indices[i] != i + n - r:
                    break
            else:
                return
            indices[i] += 1
            for j in range(i + 1, r):
                indices[j] = indices[j - 1] + 1
            yield tuple(pool[i] for i in indices)
            kmatris.append(satir[1], tuple(pool[i] for i in indices), satir[0])
            # time.sleep(3)
            # print(str(tuple(pool[i] for i in indices)))


        # Toplamlara göre kombinasyonları sırala ve en küçük k tanesini seç
        en_kucuk_kombinasyonlar = sorted(kmatris)[:k]


        yeni_satir_matrisi = []
        # Seçilen kombinasyonları yazdır
        for toplam, kombinasyon, degerler in en_kucuk_kombinasyonlar:
            # print(f"Kombinasyon: {kombinasyon}, Sütun Değerleri: {degerler}, Toplam: {toplam}")
            #print(f"Kombinasyon: {kombinasyon},  Toplam: {toplam}")
            yeni_satir_matrisi.append(degerler)
        satir = en_kucuk_sutun_degerleri_ve_toplami(yeni_satir_matrisi)
        print(satir[0])
        print(satir[1])



        #seçilen satırların matrise aktarılması
        yeni_satir_matrisi = []
        #print("Random İşlem Satırlar")
       # for kombinasyon in en_kucuk_kombinasyonlar:
            #yeni_satir_matrisi.append(item[3])
            #print(kombinasyon)
        #print(yeni_satir_matrisi)
        #matrise ait en küçük sütun değerli satırın oluşturulması
        #satir = en_kucuk_sutun_degerleri_ve_toplami(yeni_satir_matrisi)
        #print(satir[0])
        #print(satir[1])
        #print("************")

        end_time = time.time()  # İşlem bitiş zamanı
        elapsed_time = end_time - start_time  # İşlem süresi

        data_dict={'algoritma':'bruteforce','dugumsayisi':dugumsayisi,'dugumderecesi':dugumderecesi,'motiforani':motiforani,'motifsayisi':motifsayisi,'k':k,'mesafe':satir[1],'zamansn':elapsed_time}
        add_row_to_excel(file_path,data_dict)

        print("Toplam Süre: " + str(elapsed_time))
    """

    def bruteforcealg(matris, r, k):  # r: bileşke oluşturacak seçim sayısı(kombinasyonların sayısı), k: secim sayısı

        # k = 2
        kmatris = []
        # _kmatris=[]
        # combinations('ABCD', 2) → AB AC AD BC BD CD
        # combinations(range(4), 3) → 012 013 023 123
        pool = tuple(matris)
        n = len(pool)
        if r > n:
            return
        indices = list(range(r))
        # yield tuple(pool[i] for i in indices)
        # print(str(tuple(pool[i] for i in indices)))
        satir = en_kucuk_sutun_degerleri_ve_toplami(tuple(pool[i] for i in indices))
        # print(satir[0])
        # print(satir[1])
        # print(indices)

        kmatris.append((sum(satir[0]), str(indices), satir[0]))
        kmatris = sorted(kmatris)[:k]
        while True:
            for i in reversed(range(r)):
                if indices[i] != i + n - r:
                    break
            else:
                print("bitti")
                for sss in kmatris:
                    print(sss[0])
                print(len(kmatris))
                return kmatris
            indices[i] += 1
            for j in range(i + 1, r):
                indices[j] = indices[j - 1] + 1
            # yield tuple(pool[i] for i in indices)
            satir = en_kucuk_sutun_degerleri_ve_toplami(tuple(pool[i] for i in indices))
            # print(satir[0])
            # print(satir[1])
            # print(indices)
            # exit()
            # print("sonuc satırları")
            # print(indices)
            kmatris.append((sum(satir[0]), str(indices), satir[0]))
            # Toplamlara göre kombinasyonları sırala ve en küçük k tanesini seç
            kmatris = sorted(kmatris)[:k]
            # for itt in kmatris:
            #    print(itt)

        # print(kmatris)

    """
    BruteForce bitiş
    """

    # Greed1 açğrım
    print(
        "*************************************************Greedy 1 Sonuçları******************************************")
    for secim in list(secimsayisi):
        # İşlemi uygula ve sonuçları yazdır
        removed_original_rows, removed_original_indices, removed_sums, elapsed_time = greedy1alg(matrix, secim)
        for i, (row, index, sum_) in enumerate(zip(removed_original_rows, removed_original_indices, removed_sums), 1):
             #print(f"{i}. Seçilen Motif: {row}, İndeksi: {index}, Motifi: {motifler[index]}, Toplamı: {sum_}")
             print(f"{motifler[index]}")
            #print(f"{i}. Seçilen Motif, Orijinal İndeksi: {index}, Toplamı: {sum_}")
        sorted_data = sorted(zip(removed_original_rows, removed_original_indices, removed_sums), key=lambda x: x[2])

        for i, (row, index, sum_) in enumerate(sorted_data, 1):
            print(f"{motifler[index]}")



        print(f"İşlem Süresi: {elapsed_time:.6f} saniye")
    # ********************************************************************************************************
    # Greed2 çağrım
    """
    print("***********************************************Greedy 2 Sonuçları******************************************")
    for secim in list(secimsayisi):
        # k=secim
        # Fonksiyonu çağır
        greedy2alg(matrix, secim)
    """
    # ********************************************************************************************************
    print("*************************************************Random Sonuçları******************************************")
    for secim in list(secimsayisi):
        # Fonksiyonu çağır ve sonucu yazdır
        randomalg(matrix, secim, islemtekrarsayisi)

    # ********************************************************************************************************
    """
    print("********************************************Brute Force Sonuçları******************************************")
    # brute force çağrım

    for secim in list(secimsayisi):
        # Fonksiyonu çağır
        # bruteforcealg(matrix, kimbinasyonsayisi, secim)
        start_time = time.time()  # Zaman ölçümünü başlat
        sonucmat=bruteforcealg(matrix, kimbinasyonsayisi, secim)
        toplam = 0
        for satır in sonucmat:
            toplam += satır[0]
        ortalama = toplam / len(sonucmat)
        print("ortalama: "+str(ortalama))
        # Çalışma süresini yazdır
        end_time = time.time()  # Zaman ölçümünü bitir
        elapsed_time=end_time-start_time

        data_dict={'algoritma':'bruteforcealg','dugumsayisi':dugumsayisi,'dugumderecesi':dugumderecesi,'motiforani':motiforani,'motifsayisi':motifsayisi,'k':secim,'mesafe':ortalama,'zamansn':elapsed_time}
        add_row_to_excel(file_path,data_dict)
        print(f"Kodun çalışma süresi: {end_time - start_time:.6f} saniye")
    """
    """    
    for secim in list(secimsayisi):
        # Fonksiyonu çağır
        # bruteforcealg(matrix, kimbinasyonsayisi, secim)
        start_time = time.time()  # Zaman ölçümünü başlat
        sonucmat=bruteforcealg(matrix, secim, 1)    #secim kombinasyon sayısı olarak kullanılıyor bu algoritmada. Diğerinde seçilen eleman sayısı
        toplam = 0
        for satır in sonucmat:
            toplam += satır[0]
        ortalama = toplam / len(sonucmat)
        print("ortalama: "+str(ortalama))
        # Çalışma süresini yazdır
        end_time = time.time()  # Zaman ölçümünü bitir
        elapsed_time=end_time-start_time

        data_dict={'algoritma':'bruteforcealg','dugumsayisi':dugumsayisi,'dugumderecesi':dugumderecesi,'motiforani':motiforani,'motifsayisi':motifsayisi,'k':secim,'mesafe':ortalama,'zamansn':elapsed_time}
        add_row_to_excel(file_path,data_dict)
        print(f"Kodun çalışma süresi: {end_time - start_time:.6f} saniye")
    """


print("Başladı")
# Sonucu ekrana yazdır
for satir in satir_sutun_degerleri:
    # print(satir)
    # print(satir["NodePath"])
    if satir["Durum"] != 1:
        sonuclandir(satir["NodePath"], satir["MotifPath"], satir["DugumDerecesi"], satir["MotifOrani"],satir["MotifTipi"],satir["NodePath"],satir["MotifPath"])
print("Bitti")
# exit()


